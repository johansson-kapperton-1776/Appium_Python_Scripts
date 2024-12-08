import json
from appium import webdriver
from appium.options.android import UiAutomator2Options
from appium.webdriver.common.appiumby import AppiumBy
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
import openpyxl
from openpyxl.styles import Font
import logging
import time
import os
from datetime import datetime

# Configure logging
logging.basicConfig(level=logging.INFO)

# Load device configurations
with open('device_configs.json') as config_file:
    config_data = json.load(config_file)

# Define the number of trials
num_trials = 4


# Function to run tests on a device
def run_device_tests(device_config):
    logging.info(f"Starting tests on {device_config['deviceName']} with UID {device_config['deviceUID']}")

    # Excel file setup
    excel_file = f"C:\\Users\\actio\\Documents\\Fast_automation_results_andriod\\{device_config['deviceUniqueId']}_results.xlsx"

    # Load or create workbook and select the active sheet
    if os.path.exists(excel_file):
        workbook = openpyxl.load_workbook(excel_file)
        sheet = workbook.active
    else:
        workbook = openpyxl.Workbook()
        sheet = workbook.active
        sheet.title = 'Speed Test Results'
        sheet.append(
            ["Timestamp", "Download Speed (Mbps)", "Upload Speed (Mbps)", "Avg Download Speed", "Avg Upload Speed",
             "Test Run ID"])

    # Determine the next Test Run ID
    last_run_id = 0
    for value in sheet['F']:
        if value.value and value.value.startswith("TR-"):
            last_run_id = max(last_run_id, int(value.value.split('-')[1]))
    test_run_id = f"TR-{last_run_id + 1}"

    # Initialize the driver using UiAutomator2Options
    options = UiAutomator2Options()
    options.platform_name = device_config['platformName']
    options.device_name = device_config['deviceName']
    options.udid = device_config['deviceUID']
    options.app_package = device_config['appPackage']
    options.app_activity = device_config['appActivity']
    options.no_reset = True
    options.full_reset = False

    driver = webdriver.Remote("http://localhost:4723/wd/hub", options=options)
    logging.info(f"Driver initialized successfully for {device_config['deviceName']}.")

    try:
        for i in range(num_trials):  # Use the defined number of trials
            logging.info(f"Starting trial {i + 1} for {device_config['deviceName']}...")
            trial_start_time = datetime.now().strftime("%Y-%m-%d %H:%M:%S")

            # Wait for the "Refresh/Start Test" button and click it
            WebDriverWait(driver, 20).until(
                EC.presence_of_element_located(
                    (AppiumBy.XPATH, "//android.widget.TextView[@resource-id='speed-progress-indicator-icon']"))
            )
            start_button = driver.find_element(AppiumBy.XPATH, "//android.widget.TextView[@resource-id='speed-progress-indicator-icon']")
            start_button.click()

            # Sleep to allow the download speed test to complete
            time.sleep(10)  # Adjust the sleep time as necessary based on test duration

            # Wait for the download speed to be displayed and extract the value
            download_speed_element = driver.find_element(AppiumBy.XPATH, "//android.widget.TextView[@resource-id='speed-value']")
            download_speed = float(download_speed_element.text)

            # Sleep to allow the upload speed test to complete
            time.sleep(45)  # Adjust the sleep time as necessary based on test duration

            # Wait for the upload speed to be displayed and extract the value
            upload_speed_element = driver.find_element(AppiumBy.XPATH, "//android.widget.TextView[@resource-id='upload-value']")
            upload_speed = float(upload_speed_element.text)

            # Log the extracted speeds
            logging.info(f"Download Speed: {download_speed} Mbps")
            logging.info(f"Upload Speed: {upload_speed} Mbps")

            # Save the results to the Excel sheet
            sheet.append([trial_start_time, download_speed, upload_speed, None, None, test_run_id])

            # Short delay before starting the next trial
            time.sleep(5)

        # Calculate averages at the end of the sheet for each run
        last_row = sheet.max_row
        avg_download_formula = f"=AVERAGE(B{last_row - num_trials + 1}:B{last_row})"
        avg_upload_formula = f"=AVERAGE(C{last_row - num_trials + 1}:C{last_row})"
        sheet.append([None, None, None, avg_download_formula, avg_upload_formula, f"Average for {test_run_id}"])
        for col in range(4, 6):  # Apply bold formatting to average cells
            cell = sheet.cell(row=sheet.max_row, column=col)
            cell.font = Font(bold=True)

        # Leave a space after each test run
        sheet.append([None] * 6)

        workbook.save(excel_file)
        logging.info(f"Results saved to {excel_file}")

    finally:
        driver.quit()


# Iterate over each device configuration and run tests
for device in config_data['devices']:
    run_device_tests(device)
