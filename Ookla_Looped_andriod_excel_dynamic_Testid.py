from appium import webdriver
from appium.options.android import UiAutomator2Options
from appium.webdriver.common.appiumby import AppiumBy
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
import openpyxl
from openpyxl.styles import Font
import logging
import time
import os

# Configure logging
logging.basicConfig(level=logging.INFO)

# Desired capabilities using AppiumOptions
options = UiAutomator2Options()
options.platform_name = "Android"
options.device_name = "Samsung Galaxy A13"
options.app_package = "org.zwanoo.android.speedtest"
options.app_activity = "com.ookla.mobile4.screens.main.MainActivity"
options.no_reset = True
options.full_reset = False

# Excel file setup
excel_file = "C:\\Users\\actio\\Documents\\Ookla_automation_results_andriod\\A13_results_final.xlsx"

# Load or create workbook and select the active sheet
if os.path.exists(excel_file):
    workbook = openpyxl.load_workbook(excel_file)
    sheet = workbook.active
    if not sheet['E1'].value:  # Ensure there is a header for the Test Run ID
        sheet['E1'] = 'Test Run ID'
else:
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = 'Speed Test Results'
    sheet.append(["Download Speed (Mbps)", "Upload Speed (Mbps)", "Avg Download Speed", "Avg Upload Speed", "Test Run ID"])

# Define the number of trials
num_trials = 5

# Determine the next Test Run ID
last_run_id = 0
for value in sheet['E']:
    if value.value and value.value.startswith("TR-"):
        last_run_id = max(last_run_id, int(value.value.split('-')[1]))
test_run_id = f"TR-{last_run_id + 1}"

# Initialize the driver
logging.info("Initializing the driver...")
driver = None

try:
    driver = webdriver.Remote("http://localhost:4723/wd/hub", options=options)
    logging.info("Driver initialized successfully.")

    for i in range(num_trials):
        logging.info(f"Starting trial {i+1}...")

        # Wait for the "GO" button and click
        WebDriverWait(driver, 20).until(
            EC.presence_of_element_located((AppiumBy.ID, "org.zwanoo.android.speedtest:id/go_button"))
        )
        go_button = driver.find_element(AppiumBy.ID, "org.zwanoo.android.speedtest:id/go_button")
        go_button.click()

        # Wait for results
        WebDriverWait(driver, 60).until(
            EC.presence_of_element_located(
                (AppiumBy.XPATH, '//android.widget.FrameLayout[@content-desc="DOWNLOAD"]/android.view.ViewGroup'))
        )

        download_speed_element = WebDriverWait(driver, 45).until(
            EC.presence_of_element_located((AppiumBy.XPATH, '//android.widget.FrameLayout[@content-desc="DOWNLOAD"]/android.view.ViewGroup/android.widget.TextView[@resource-id="org.zwanoo.android.speedtest:id/txt_test_result_value"]'))
        )
        upload_speed_element = WebDriverWait(driver, 45).until(
            EC.presence_of_element_located((AppiumBy.XPATH, '//android.widget.FrameLayout[@content-desc="UPLOAD"]/android.view.ViewGroup/android.widget.TextView[@resource-id="org.zwanoo.android.speedtest:id/txt_test_result_value"]'))
        )

        download_speed = float(download_speed_element.text.split()[0])
        upload_speed = float(upload_speed_element.text.split()[0])

        logging.info(f"Download Speed: {download_speed} Mbps")
        logging.info(f"Upload Speed: {upload_speed} Mbps")

        sheet.append([download_speed, upload_speed, None, None, test_run_id])

        time.sleep(2)
        close_icon = WebDriverWait(driver, 10).until(
            EC.element_to_be_clickable((AppiumBy.ID, "org.zwanoo.android.speedtest:id/closeIcon"))
        )
        close_icon.click()
        time.sleep(3)

    last_row = sheet.max_row
    avg_download_formula = f"=AVERAGE(A2:A{last_row})"
    avg_upload_formula = f"=AVERAGE(B2:B{last_row})"
    sheet.cell(row=last_row, column=3, value=avg_download_formula)
    sheet.cell(row=last_row, column=4, value=avg_upload_formula)

    for col in range(3, 5):
        cell = sheet.cell(row=last_row, column=col)
        cell.font = Font(bold=True)

    workbook.save(excel_file)
    logging.info(f"Results saved to {excel_file}")

finally:
    if driver:
        driver.quit()
