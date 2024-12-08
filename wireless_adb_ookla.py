import json
from appium import webdriver
from appium.options.android import UiAutomator2Options
from appium.webdriver.common.appiumby import AppiumBy
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
import openpyxl
from openpyxl.styles import Font
import logging
import time
import os
from datetime import datetime

# Configure logging
logging.basicConfig(level=logging.INFO)

# Load device configurations
with open('adb_devices.json') as config_file:
    config_data = json.load(config_file)

# Define the number of trials
num_trials = 4


# Function to run tests on a device
def run_device_tests(device_config):
    logging.info(f"Starting tests on {device_config['deviceName']} with UID {device_config['deviceUID']}")

    # Excel file setup
    excel_file = f"C:\\Users\\actio\\Documents\\Ookla_automation_results_andriod\\{device_config['deviceUniqueId']}_results.xlsx"

    # Load or create workbook and select the active sheet
    if os.path.exists(excel_file):
        workbook = openpyxl.load_workbook(excel_file)
        sheet = workbook.active
    else:
        workbook = openpyxl.Workbook()
        sheet = workbook.active
        sheet.title = 'Speed Test Results'
        sheet.append(["Timestamp", "Download Speed (Mbps)", "Upload Speed (Mbps)", "Avg Download Speed", "Avg Upload Speed", "Test Run ID"])

    # Determine the next Test Run ID
    last_run_id = 0
    for value in sheet['F']:
        if value.value and value.value.startswith("TR-"):
            last_run_id = max(last_run_id, int(value.value.split('-')[1]))
    test_run_id = f"TR-{last_run_id + 1}"

    # Get IP address for wireless ADB connection
    ip_address = device_config.get('ipAddress')

    # Connect to the device via wireless ADB
    os.system(f"adb connect {ip_address}:5555")

    # Initialize the driver using UiAutomator2Options
    options = UiAutomator2Options()
    options.platform_name = device_config['platformName']
    options.device_name = device_config['deviceName']
    options.udid = device_config['deviceUID']
    options.app_package = device_config['appPackage']
    options.app_activity = device_config['appActivity']
    options.no_reset = True
    options.full_reset = False

    driver = webdriver.Remote("http://localhost:4723/wd/hub", options=options)
    logging.info(f"Driver initialized successfully for {device_config['deviceName']}.")

    try:
        for i in range(num_trials):  # Use the defined number of trials
            logging.info(f"Starting trial {i + 1} for {device_config['deviceName']}...")
            trial_start_time = datetime.now().strftime("%Y-%m-%d %H:%M:%S")

            # Wait for the "GO" button and click
            WebDriverWait(driver, 20).until(
                EC.presence_of_element_located((AppiumBy.ID, "org.zwanoo.android.speedtest:id/go_button"))
            )
            go_button = driver.find_element(AppiumBy.ID, "org.zwanoo.android.speedtest:id/go_button")
            go_button.click()

            # Wait for results
            WebDriverWait(driver, 60).until(
                EC.presence_of_element_located((AppiumBy.XPATH, '//android.widget.FrameLayout[@content-desc="DOWNLOAD"]/android.view.ViewGroup'))
            )

            download_speed_element = WebDriverWait(driver, 45).until(
                EC.presence_of_element_located((AppiumBy.XPATH, '//android.widget.FrameLayout[@content-desc="DOWNLOAD"]/android.view.ViewGroup/android.widget.TextView[@resource-id="org.zwanoo.android.speedtest:id/txt_test_result_value"]'))
            )
            upload_speed_element = WebDriverWait(driver, 45).until(
                EC.presence_of_element_located((AppiumBy.XPATH, '//android.widget.FrameLayout[@content-desc="UPLOAD"]/android.view.ViewGroup/android.widget.TextView[@resource-id="org.zwanoo.android.speedtest:id/txt_test_result_value"]'))
            )

            download_speed = float(download_speed_element.text.split()[0])
            upload_speed = float(upload_speed_element.text.split()[0])

            logging.info(f"Download Speed: {download_speed} Mbps")
            logging.info(f"Upload Speed: {upload_speed} Mbps")

            sheet.append([trial_start_time, download_speed, upload_speed, None, None, test_run_id])

            time.sleep(2)
            close_icon = WebDriverWait(driver, 10).until(
                EC.element_to_be_clickable((AppiumBy.ID, "org.zwanoo.android.speedtest:id/closeIcon"))
            )
            close_icon.click()
            time.sleep(3)

        # Calculate averages at the end of the sheet
        last_row = sheet.max_row
        avg_download_formula = f"=AVERAGE(B2:B{last_row})"
        avg_upload_formula = f"=AVERAGE(C2:C{last_row})"
        sheet.cell(row=last_row + 1, column=4, value=avg_download_formula)
        sheet.cell(row=last_row + 1, column=5, value=avg_upload_formula)
        for col in range(4, 6):  # Apply bold formatting to average cells
            cell = sheet.cell(row=last_row + 1, column=col)
            cell.font = Font(bold=True)

        workbook.save(excel_file)
        logging.info(f"Results saved to {excel_file}")

    finally:
        driver.quit()
        # Disconnect wireless ADB connection
        os.system(f"adb disconnect {ip_address}:5555")


# Iterate over each device configuration and run tests
for device in config_data['devices']:
    run_device_tests(device)